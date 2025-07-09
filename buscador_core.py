import os
import io
import fitz  # PyMuPDF
import docx  # python-docx
import openpyxl
from openpyxl.utils import get_column_letter
import xlrd

def _get_context_snippets(full_text, search_string, context_chars):
    """
    Función auxiliar que encuentra ocurrencias y devuelve fragmentos de contexto.
    Ahora maneja el caso donde context_chars es 0.
    """
    snippets = []
    text_lower = full_text.lower()
    search_string_lower = search_string.lower()
    last_pos = 0
    
    while True:
        pos = text_lower.find(search_string_lower, last_pos)
        if pos == -1:
            break
        
        found_text = full_text[pos:pos + len(search_string)]
        
        # === CAMBIO AQUÍ: Formato condicional basado en context_chars ===
        if context_chars > 0:
            start_context = max(0, pos - context_chars)
            end_context = min(len(full_text), pos + len(search_string) + context_chars)
            pre_context = full_text[start_context:pos].replace('\n', ' ').replace('\r', '')
            post_context = full_text[pos + len(search_string):end_context].replace('\n', ' ').replace('\r', '')
            snippet = f"  └─ Contexto: ...{pre_context} >>>{found_text}<<< {post_context}..."
        else: # Si el contexto es 0, mostrar solo la ocurrencia
            snippet = f"  └─ Ocurrencia exacta: >>>{found_text}<<<"

        snippets.append(snippet)
        last_pos = pos + 1
        
    return snippets

# El resto de las funciones procesar_* y generar_informe se mantienen
# casi idénticas, ya que la lógica principal del cambio está encapsulada
# en _get_context_snippets. Solo paso los parámetros.

def procesar_pdf(ruta_archivo, lista_strings, context_chars):
    hallazgos, problemas = [], []
    nombre_base = os.path.basename(ruta_archivo)
    try:
        documento = fitz.open(ruta_archivo)
        if not documento.page_count:
            problemas.append(f"Archivo: '{nombre_base}' -> ERROR: El PDF está vacío o corrupto.")
            return hallazgos, problemas
        
        for num_pagina, pagina in enumerate(documento, start=1):
            page_text = pagina.get_text("text")
            if not page_text.strip(): continue
            
            for string_a_buscar in lista_strings:
                snippets = _get_context_snippets(page_text, string_a_buscar, context_chars)
                if snippets:
                    hallazgos.append(f"\nArchivo: '{nombre_base}', Página: {num_pagina} -> Encontrado: '{string_a_buscar}'")
                    hallazgos.extend(snippets)
        documento.close()
    except Exception as e:
        problemas.append(f"Archivo: '{nombre_base}' -> ERROR: No se pudo procesar. Razón: {e}")
    return hallazgos, problemas

def procesar_docx(ruta_archivo, lista_strings, context_chars):
    # ... (El código de esta función no cambia, solo recibe context_chars y lo pasa)
    hallazgos, problemas = [], []
    nombre_base = os.path.basename(ruta_archivo)
    try:
        documento = docx.Document(ruta_archivo)
        for num_parrafo, parrafo in enumerate(documento.paragraphs, start=1):
            if not parrafo.text.strip(): continue
            for string_a_buscar in lista_strings:
                snippets = _get_context_snippets(parrafo.text, string_a_buscar, context_chars)
                if snippets:
                    hallazgos.append(f"\nArchivo: '{nombre_base}', Párrafo: {num_parrafo} -> Encontrado: '{string_a_buscar}'")
                    hallazgos.extend(snippets)
    except Exception as e:
        problemas.append(f"Archivo: '{nombre_base}' -> ERROR: No se pudo procesar. Razón: {e}")
    return hallazgos, problemas

def procesar_excel(ruta_archivo, lista_strings, context_chars):
    # ... (El código de esta función no cambia, solo recibe context_chars y lo pasa)
    hallazgos, problemas = [], []
    nombre_base = os.path.basename(ruta_archivo)
    try:
        if ruta_archivo.lower().endswith('.xlsx'):
            wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
            for nombre_hoja in wb.sheetnames:
                hoja = wb[nombre_hoja]
                for fila_idx, fila in enumerate(hoja.iter_rows(), start=1):
                    for col_idx, celda in enumerate(fila, start=1):
                        if celda.value and isinstance(celda.value, str):
                            for string_a_buscar in lista_strings:
                                snippets = _get_context_snippets(str(celda.value), string_a_buscar, context_chars)
                                if snippets:
                                    celda_ref = f"{get_column_letter(col_idx)}{fila_idx}"
                                    hallazgos.append(f"\nArchivo: '{nombre_base}', Hoja: '{nombre_hoja}', Celda: {celda_ref} -> Encontrado: '{string_a_buscar}'")
                                    hallazgos.extend(snippets)
        elif ruta_archivo.lower().endswith('.xls'):
            wb = xlrd.open_workbook(ruta_archivo)
            for nombre_hoja in wb.sheet_names():
                hoja = wb.sheet_by_name(nombre_hoja)
                for fila_idx in range(hoja.nrows):
                    for col_idx in range(hoja.ncols):
                        valor_celda = str(hoja.cell_value(fila_idx, col_idx))
                        if not valor_celda.strip(): continue
                        for string_a_buscar in lista_strings:
                            snippets = _get_context_snippets(valor_celda, string_a_buscar, context_chars)
                            if snippets:
                                celda_ref = f"{chr(65+col_idx)}{fila_idx+1}"
                                hallazgos.append(f"\nArchivo: '{nombre_base}', Hoja: '{nombre_hoja}', Celda: ~{celda_ref} -> Encontrado: '{string_a_buscar}'")
                                hallazgos.extend(snippets)
    except Exception as e:
        problemas.append(f"Archivo: '{nombre_base}' -> ERROR: No se pudo procesar. Razón: {e}")
    return hallazgos, problemas

def procesar_txt(ruta_archivo, lista_strings, context_chars):
    # ... (El código de esta función no cambia, solo recibe context_chars y lo pasa)
    hallazgos, problemas = [], []
    nombre_base = os.path.basename(ruta_archivo)
    try:
        with open(ruta_archivo, 'r', encoding='utf-8', errors='ignore') as f:
            for num_linea, linea in enumerate(f, start=1):
                if not linea.strip(): continue
                for string_a_buscar in lista_strings:
                    snippets = _get_context_snippets(linea, string_a_buscar, context_chars)
                    if snippets:
                        hallazgos.append(f"\nArchivo: '{nombre_base}', Línea: {num_linea} -> Encontrado: '{string_a_buscar}'")
                        hallazgos.extend(snippets)
    except Exception as e:
        problemas.append(f"Archivo: '{nombre_base}' -> ERROR: No se pudo procesar. Razón: {e}")
    return hallazgos, problemas

def generar_informe(carpeta_entrada, lista_strings, context_chars=240):
    # La lógica de esta función para procesar archivos y generar las secciones del
    # informe es exactamente la misma que en la versión anterior.
    # Solo se asegura de pasar 'context_chars' a todas las funciones 'procesar_*'.
    
    hallazgos_totales, archivos_problematicos, archivos_ignorados = [], [], []
    archivos_procesados = 0
    set_archivos_con_problemas = set()
    extensiones_soportadas = ('.pdf', '.docx', '.xlsx', '.xls', '.txt')

    for nombre_archivo in os.listdir(carpeta_entrada):
        ruta_completa = os.path.join(carpeta_entrada, nombre_archivo)
        if os.path.isfile(ruta_completa):
            extension = os.path.splitext(nombre_archivo)[1].lower()
            if extension in extensiones_soportadas:
                archivos_procesados += 1
                if extension == '.pdf': hallazgos, problemas = procesar_pdf(ruta_completa, lista_strings, context_chars)
                elif extension == '.docx': hallazgos, problemas = procesar_docx(ruta_completa, lista_strings, context_chars)
                elif extension in ['.xlsx', '.xls']: hallazgos, problemas = procesar_excel(ruta_completa, lista_strings, context_chars)
                elif extension == '.txt': hallazgos, problemas = procesar_txt(ruta_completa, lista_strings, context_chars)
                
                hallazgos_totales.extend(hallazgos)
                archivos_problematicos.extend(problemas)
                if problemas:
                    set_archivos_con_problemas.add(nombre_archivo)
            else:
                archivos_ignorados.append(nombre_archivo)
    
    total_con_problemas = len(set_archivos_con_problemas)
    total_ignorados = len(archivos_ignorados)
    total_sin_problemas = archivos_procesados - total_con_problemas
    total_seleccionados = archivos_procesados + total_ignorados

    output = io.StringIO()
    output.write("="*30 + " INFORME DE BÚSQUEDA " + "="*30 + "\n")
    output.write(f"Textos Buscados: {lista_strings}\n")
    output.write(f"Cantidad de Caracteres de Contexto anteriores y posteriores al texto hallado: {context_chars}\n")
    output.write(f"Extensiones Soportadas: {', '.join(extensiones_soportadas)}\n")
    output.write("="*79 + "\n")
    
    output.write("\n--- OCURRENCIAS HALLADAS ---\n")
    if hallazgos_totales:
        for hallazgo in hallazgos_totales:
            output.write(hallazgo + "\n")
    else:
        output.write("No se encontraron ocurrencias de los textos buscados.\n")

    output.write("\n\n--- ARCHIVOS PROCESADOS CON PROBLEMAS O ADVERTENCIAS ---\n")
    if archivos_problematicos:
        for problema in archivos_problematicos:
            output.write(problema + "\n")
    else:
        output.write("Todos los archivos soportados fueron analizados sin errores.\n")

    output.write("\n\n--- ARCHIVOS NO SOPORTADOS E IGNORADOS ---\n")
    output.write(f"Total: {total_ignorados}\n\n")
    if archivos_ignorados:
        for archivo in sorted(archivos_ignorados):
            output.write(f"- {archivo}\n")
    else:
        output.write("No se encontraron archivos con formatos no soportados.\n")
    
    output.write("\n\n" + "="*33 + " RESUMEN FINAL " + "="*33 + "\n")
    output.write(f"TOTAL DE ARCHIVOS SELECCIONADOS: {total_seleccionados}\n")
    output.write(f"  - TOTAL DE ARCHIVOS PROCESADOS SIN PROBLEMAS: {total_sin_problemas}\n")
    output.write(f"  - TOTAL DE ARCHIVOS PROCESADOS CON PROBLEMAS O ADVERTENCIAS: {total_con_problemas}\n")
    output.write(f"  - TOTAL DE ARCHIVOS NO SOPORTADOS E IGNORADOS: {total_ignorados}\n")
    
    return output.getvalue()